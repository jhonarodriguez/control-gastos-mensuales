import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json

class GeneradorExcelDrive:
    """
    Generador de Excel optimizado para Google Drive
    Crea un archivo único con hojas mensuales
    """
    
    def __init__(self, config_path='config/configuracion.json'):
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = json.load(f)
        
        self.colores = {
            'encabezado': '366092',
            'subencabezado': '4472C4',
            'ingreso': 'C6EFCE',
            'gasto': 'FFC7CE',
            'total': 'FFEB9C',
            'banco': 'E8F4FD',
            'diferencia_positiva': 'D4EDDA',
            'diferencia_negativa': 'F8D7DA',
            'neutro': 'D9D9D9'
        }
        
        self.bordes = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def crear_excel_drive(self, mes_nombre=None, anio=None):
        """
        Crea un nuevo Excel para Drive con hoja mensual
        """
        if mes_nombre is None:
            mes_nombre = self._obtener_nombre_mes_actual()
        if anio is None:
            anio = datetime.now().year
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Crear hoja del mes actual
        ws_mes = wb.create_sheet(mes_nombre)
        self._crear_hoja_mes(ws_mes, mes_nombre, anio)
        
        # Crear hoja de histórico
        ws_historico = wb.create_sheet('Histórico')
        self._crear_hoja_historico(ws_historico)
        
        return wb
    
    def _obtener_nombre_mes_actual(self):
        """Obtiene el nombre del mes actual en español"""
        meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        return meses[datetime.now().month - 1]
    
    def _crear_hoja_mes(self, ws, mes_nombre, anio):
        """Crea la hoja de un mes específico con saldo bancario"""
        
        # Título principal
        ws['A1'] = f'CONTROL DE GASTOS - {mes_nombre.upper()} {anio}'
        ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colores['encabezado'], 
                                     end_color=self.colores['encabezado'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 35
        
        # SECCIÓN 1: SALDO BANCARIO
        fila = 3
        ws[f'A{fila}'] = 'SALDO BANCARIO'
        ws[f'A{fila}'].font = Font(size=14, bold=True, color='FFFFFF')
        ws[f'A{fila}'].fill = PatternFill(start_color='845EF7', end_color='845EF7', fill_type='solid')
        ws.merge_cells(f'A{fila}:F{fila}')
        
        fila += 1
        # Encabezados
        headers = ['Concepto', 'Monto (COP)', 'Diferencia', 'Estado', 'Fecha', 'Notas']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=fila, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colores['subencabezado'], 
                                   end_color=self.colores['subencabezado'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.bordes
        
        fila += 1
        saldo_mes_pasado = self.config.get('saldo_mes_anterior', 0)
        saldo_actual_config = self.config.get('saldo_bancario', {}).get('valor_actual', 0)
        
        # Saldo mes pasado (al inicio del mes)
        ws[f'A{fila}'] = 'Saldo Inicio Mes (Mes Pasado)'
        ws[f'B{fila}'] = saldo_mes_pasado
        ws[f'B{fila}'].number_format = '$#,##0.00'
        ws[f'B{fila}'].fill = PatternFill(start_color=self.colores['banco'], 
                                         end_color=self.colores['banco'], fill_type='solid')
        for col in range(1, 7):
            ws.cell(row=fila, column=col).border = self.bordes
        
        fila += 1
        # Ingresos del mes
        ws[f'A{fila}'] = 'Ingresos del Mes'
        ws[f'B{fila}'] = self.config['sueldo']['valor_fijo']
        ws[f'B{fila}'].number_format = '$#,##0.00'
        ws[f'B{fila}'].fill = PatternFill(start_color=self.colores['ingreso'], 
                                         end_color=self.colores['ingreso'], fill_type='solid')
        for col in range(1, 7):
            ws.cell(row=fila, column=col).border = self.bordes
        
        fila += 1
        # Total Gastos del Mes
        total_gastos = sum(gasto.get('valor', 0) for gasto in self.config['gastos_fijos'].values())
        ws[f'A{fila}'] = 'Total Gastos del Mes'
        ws[f'B{fila}'] = total_gastos
        ws[f'B{fila}'].number_format = '$#,##0.00'
        ws[f'B{fila}'].fill = PatternFill(start_color=self.colores['gasto'], 
                                         end_color=self.colores['gasto'], fill_type='solid')
        for col in range(1, 7):
            ws.cell(row=fila, column=col).border = self.bordes
        
        fila += 1
        # Saldo Calculado (lo que debería haber)
        fila_saldo_calculado = fila
        ws[f'A{fila}'] = 'Saldo Calculado (Debería tener)'
        ws[f'A{fila}'].font = Font(bold=True)
        ws[f'B{fila}'] = f'=B{fila-3}+B{fila-2}-B{fila-1}'
        ws[f'B{fila}'].number_format = '$#,##0.00'
        ws[f'B{fila}'].font = Font(bold=True)
        ws[f'B{fila}'].fill = PatternFill(start_color=self.colores['total'], 
                                         end_color=self.colores['total'], fill_type='solid')
        for col in range(1, 7):
            ws.cell(row=fila, column=col).border = self.bordes
            ws.cell(row=fila, column=col).fill = PatternFill(start_color=self.colores['total'], 
                                                             end_color=self.colores['total'], fill_type='solid')
        
        fila += 1
        # Saldo Real en Banco (actualizado manualmente)
        fila_saldo_real = fila
        ws[f'A{fila}'] = 'Saldo REAL en Banco'
        ws[f'A{fila}'].font = Font(bold=True, size=11)
        ws[f'B{fila}'] = saldo_actual_config
        ws[f'B{fila}'].number_format = '$#,##0.00'
        ws[f'B{fila}'].font = Font(bold=True, size=11)
        ws[f'B{fila}'].fill = PatternFill(start_color='FFD700', 
                                         end_color='FFD700', fill_type='solid')
        ws[f'F{fila}'] = 'ACTUALIZAR ESTE VALOR EN LA WEB'
        ws[f'F{fila}'].font = Font(italic=True, color='FF0000', size=9)
        for col in range(1, 7):
            ws.cell(row=fila, column=col).border = self.bordes
        
        fila += 1
        # Diferencia
        fila_diferencia = fila
        ws[f'A{fila}'] = 'DIFERENCIA'
        ws[f'A{fila}'].font = Font(bold=True, size=12, color='FFFFFF')
        ws[f'A{fila}'].fill = PatternFill(start_color='FF0000', 
                                         end_color='FF0000', fill_type='solid')
        ws[f'B{fila}'] = f'=B{fila_saldo_real}-B{fila_saldo_calculado}'
        ws[f'B{fila}'].number_format = '$#,##0.00'
        ws[f'B{fila}'].font = Font(bold=True, size=12)
        
        # Fórmula condicional para colorear según si hay diferencia
        # Nota: Las fórmulas condicionales se aplican en Excel, aquí solo ponemos la fórmula
        ws[f'C{fila}'] = f'=IF(B{fila}=0,"CUADRADO",IF(B{fila}>0,"SOBRANTE","FALTANTE"))'
        ws[f'C{fila}'].font = Font(bold=True)
        ws[f'C{fila}'].alignment = Alignment(horizontal='center')
        
        for col in range(1, 7):
            ws.cell(row=fila, column=col).border = self.bordes
            ws.cell(row=fila, column=col).fill = PatternFill(start_color='FFEB9C', 
                                                             end_color='FFEB9C', fill_type='solid')
        
        # SECCIÓN 2: DETALLE DE GASTOS
        fila += 2
        ws[f'A{fila}'] = 'DETALLE DE GASTOS'
        ws[f'A{fila}'].font = Font(size=14, bold=True, color='FFFFFF')
        ws[f'A{fila}'].fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        ws.merge_cells(f'A{fila}:F{fila}')
        
        fila += 1
        # Tabla de gastos
        gastos_headers = ['Fecha', 'Concepto', 'Categoría', 'Monto (COP)', 'Método', 'Notas']
        for col_idx, header in enumerate(gastos_headers, 1):
            cell = ws.cell(row=fila, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colores['subencabezado'], 
                                   end_color=self.colores['subencabezado'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.bordes
        
        # Fila para gastos variables (inicialmente vacía)
        fila += 1
        for col in range(1, 7):
            ws.cell(row=fila, column=col).border = self.bordes
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 40
        
        # Congelar paneles en la tabla de gastos
        ws.freeze_panes = f'A{fila-1}'
    
    def _crear_hoja_historico(self, ws):
        """Crea la hoja de histórico"""
        ws['A1'] = 'HISTÓRICO DE SALDOS BANCARIOS'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colores['encabezado'], 
                                     end_color=self.colores['encabezado'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        
        headers = ['Mes', 'Saldo Inicial', 'Ingresos', 'Gastos', 'Saldo Calculado', 'Saldo Real']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colores['subencabezado'], 
                                   end_color=self.colores['subencabezado'], fill_type='solid')
            cell.border = self.bordes
        
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col_letter].width = 18
    
    def agregar_hoja_mes(self, wb, mes_nombre, anio):
        """Agrega una nueva hoja de mes a un workbook existente"""
        if mes_nombre in wb.sheetnames:
            return wb[mes_nombre]
        
        ws = wb.create_sheet(mes_nombre)
        self._crear_hoja_mes(ws, mes_nombre, anio)
        return ws
    
    def guardar_excel_temporal(self, wb, nombre='ControlDeGastos.xlsx'):
        """Guarda el Excel en archivo temporal para subir a Drive"""
        import tempfile
        import os
        
        # Crear directorio temporal si no existe
        temp_dir = os.path.join(tempfile.gettempdir(), 'control_gastos')
        os.makedirs(temp_dir, exist_ok=True)
        
        ruta = os.path.join(temp_dir, nombre)
        
        # Cerrar cualquier libro abierto previamente
        try:
            wb.save(ruta)
        except PermissionError:
            # Si hay error de permisos, usar nombre alternativo
            import time
            ruta_alt = os.path.join(temp_dir, f"ControlDeGastos_{int(time.time())}.xlsx")
            wb.save(ruta_alt)
            return ruta_alt
        
        return ruta

if __name__ == '__main__':
    generador = GeneradorExcelDrive()
    wb = generador.crear_excel_drive()
    ruta = generador.guardar_excel_temporal(wb)
    print(f'Excel creado: {ruta}')
