import json
import os
import re
import tempfile
from datetime import datetime
from typing import Dict, List, Tuple


class ProcesadorMensajes:
    def __init__(self, config_path='config/configuracion.json'):
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = json.load(f)

        self.categorias = self.config.get('categorias_gastos', [])
        self.palabras_clave = self._crear_diccionario_palabras_clave()
        self.regex_monto = re.compile(
            r'(?<!\w)(?P<amount>\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{1,2})?|\d+)(?:\s*(?P<suffix>mil|k))?(?!\w)',
            re.IGNORECASE,
        )

    def _crear_diccionario_palabras_clave(self) -> Dict[str, str]:
        return {
            'arriendo': 'arriendo',
            'alquiler': 'arriendo',
            'mercado': 'mercado_primera_quincena',
            'supermercado': 'mercado_primera_quincena',
            'almuerzo': 'alimentacion',
            'desayuno': 'alimentacion',
            'cena': 'alimentacion',
            'cafe': 'alimentacion',
            'comida': 'alimentacion',
            'pan': 'alimentacion',
            'leche': 'alimentacion',
            'frutas': 'alimentacion',
            'gas': 'servicio_gas',
            'descuento': 'descuento_quincenal',
            'retiro': 'descuento_quincenal',
            'gimnasio': 'gimnasio',
            'gym': 'gimnasio',
            'netflix': 'netflix',
            'movistar': 'movistar',
            'youtube': 'youtube_premium',
            'google drive': 'google_drive',
            'drive': 'google_drive',
            'gamepass': 'gamepass',
            'xbox': 'gamepass',
            'mercadolibre': 'mercadolibre',
            'mercado libre': 'mercadolibre',
            'transporte': 'transporte',
            'uber': 'transporte',
            'taxi': 'transporte',
            'gasolina': 'transporte',
            'salud': 'salud_bienestar',
            'medico': 'salud_bienestar',
            'medicina': 'salud_bienestar',
            'farmacia': 'salud_bienestar',
            'entretenimiento': 'entretenimiento',
            'cine': 'entretenimiento',
            'restaurante': 'entretenimiento',
            'tecnologia': 'tecnologia',
            'celular': 'tecnologia',
            'computador': 'tecnologia',
            'educacion': 'educacion',
            'curso': 'educacion',
            'libro': 'educacion',
        }

    def _normalizar_numero(self, numero_str: str, suffix: str = '') -> float:
        txt = (numero_str or '').strip()
        if not txt:
            return 0.0

        if suffix and suffix.lower() in ('mil', 'k'):
            base = txt.replace('.', '').replace(',', '')
            try:
                return float(base) * 1000
            except ValueError:
                return 0.0

        if ',' in txt and '.' in txt:
            if txt.rfind(',') > txt.rfind('.'):
                txt = txt.replace('.', '').replace(',', '.')
            else:
                txt = txt.replace(',', '')
        elif ',' in txt:
            partes = txt.split(',')
            if len(partes[-1]) == 2:
                txt = txt.replace(',', '.')
            else:
                txt = txt.replace(',', '')
        elif '.' in txt:
            partes = txt.split('.')
            if len(partes[-1]) != 2:
                txt = txt.replace('.', '')

        try:
            return float(txt)
        except ValueError:
            return 0.0

    def _extraer_montos_con_posiciones(self, mensaje: str) -> List[Dict]:
        resultados = []
        txt = mensaje.strip()
        if not txt:
            return resultados

        for match in self.regex_monto.finditer(txt):
            start = match.start()
            end = match.end()

            # Evitar interpretar partes de fechas tipo 15/02/2026 como gastos
            prev_char = txt[start - 1] if start > 0 else ''
            next_char = txt[end] if end < len(txt) else ''
            if prev_char == '/' or next_char == '/':
                continue

            amount = self._normalizar_numero(match.group('amount'), match.group('suffix') or '')
            if amount <= 0:
                continue

            resultados.append({
                'monto': amount,
                'start': start,
                'end': end,
                'raw': match.group(0),
            })

        return resultados

    def extraer_monto(self, mensaje: str) -> float:
        montos = self._extraer_montos_con_posiciones(mensaje)
        return montos[0]['monto'] if montos else 0.0

    def _buscar_concepto_clave(self, txt: str) -> str:
        for palabra, concepto in sorted(self.palabras_clave.items(), key=lambda x: len(x[0]), reverse=True):
            patron = r'(?<!\w)' + re.escape(palabra) + r'(?!\w)'
            if re.search(patron, txt):
                return concepto
        return ''

    def detectar_categoria(self, mensaje: str) -> Tuple[str, float]:
        txt = mensaje.lower()
        concepto = self._buscar_concepto_clave(txt)
        if concepto:
            if concepto in self.config.get('gastos_fijos', {}):
                gasto_cfg = self.config['gastos_fijos'][concepto]
                categoria = gasto_cfg.get('categoria', 'Gastos Fijos')
                categoria_slug = categoria.lower().replace('/', ' ').replace('-', ' ').replace(' ', '_')
                return categoria_slug, gasto_cfg.get('valor', 0)
            return concepto, 0.0
        return 'gastos_generales', 0.0

    def _detectar_metodo_pago(self, txt: str) -> str:
        base = txt.lower()
        if any(k in base for k in ['tarjeta', 'credito', 'debito', 'visa', 'mastercard']):
            return 'Tarjeta'
        if any(k in base for k in ['transferencia', 'nequi', 'daviplata', 'pse']):
            return 'Transferencia'
        return 'Efectivo'

    def _limpiar_concepto(self, txt: str) -> str:
        if not txt:
            return ''

        cleaned = txt.lower().strip()
        cleaned = re.sub(r'[\(\)\[\]\{\}\|]', ' ', cleaned)
        cleaned = re.sub(r'\s+', ' ', cleaned).strip(' ,;:-.')
        cleaned = re.sub(
            r'^(hoy|ayer|anoche|gaste|gasté|compre|compré|pague|pagué|pago|fueron|fue|en|de|del|por|para|al|a|y|un|una|el|la|los|las)\s+',
            '',
            cleaned,
        )
        cleaned = re.sub(r'\s+', ' ', cleaned).strip(' ,;:-.')
        return cleaned

    def _es_fragmento_metodo_pago(self, txt: str) -> bool:
        base = txt.lower().strip()
        if not base:
            return False
        base = re.sub(r'^(con|en)\s+', '', base)
        return base in ('tarjeta', 'efectivo', 'transferencia', 'credito', 'debito')

    def _inferir_concepto(self, left_ctx: str, right_ctx: str, monto_raw: str) -> str:
        right_original = right_ctx.strip().lower()
        right = right_ctx.strip()
        right = re.sub(r'^(en|de|del|por|para|al|a|y)\s+', '', right, flags=re.IGNORECASE).strip()
        right = re.split(r'[,\n;]', right)[0].strip()
        right = re.sub(r'\b(y|e)\b\s*$', '', right, flags=re.IGNORECASE).strip()
        right = self._limpiar_concepto(right)

        left = left_ctx.strip()
        left = re.split(r'[,\n;]', left)[-1].strip()
        left = re.sub(r'^(y|e)\s+', '', left, flags=re.IGNORECASE).strip()
        left = self._limpiar_concepto(left)

        if right_original.startswith('y ') and left:
            return left

        if self._es_fragmento_metodo_pago(right):
            right = ''

        if right and any(ch.isalpha() for ch in right):
            return right
        if left and any(ch.isalpha() for ch in left):
            return left

        fallback = self._limpiar_concepto(left_ctx + ' ' + right_ctx)
        if fallback:
            return fallback

        return f'gasto {monto_raw}'.strip()

    def _construir_resultado(self, concepto: str, monto: float, contexto: str, metodo_pago: str, notas: str) -> Dict:
        resultado = {
            'tipo': 'gasto',
            'categoria': '',
            'concepto': concepto if concepto else 'Gasto general',
            'monto': monto,
            'fecha': datetime.now().strftime('%Y-%m-%d'),
            'metodo_pago': metodo_pago,
            'notas': notas,
            'es_gasto_fijo': False,
        }

        tipo_categoria, valor_fijo = self.detectar_categoria(concepto)
        resultado['categoria'] = tipo_categoria.replace('_', ' ').title()
        if valor_fijo > 0:
            resultado['es_gasto_fijo'] = True

        return resultado

    def procesar_mensaje_multiple(self, mensaje: str) -> List[Dict]:
        txt = (mensaje or '').strip()
        if not txt:
            return []

        montos = self._extraer_montos_con_posiciones(txt)
        if not montos:
            return []

        metodo_global = self._detectar_metodo_pago(txt)
        resultados = []

        for idx, item in enumerate(montos):
            prev_end = montos[idx - 1]['end'] if idx > 0 else 0
            next_start = montos[idx + 1]['start'] if idx + 1 < len(montos) else len(txt)

            left_ctx = txt[prev_end:item['start']]
            right_ctx = txt[item['end']:next_start]
            contexto_local = f'{left_ctx} {right_ctx}'.strip()

            metodo_local = self._detectar_metodo_pago(contexto_local)
            metodo = metodo_local if metodo_local != 'Efectivo' else metodo_global

            concepto = self._inferir_concepto(left_ctx, right_ctx, item['raw'])
            gasto = self._construir_resultado(
                concepto=concepto,
                monto=item['monto'],
                contexto=contexto_local,
                metodo_pago=metodo,
                notas=contexto_local if contexto_local else txt,
            )
            resultados.append(gasto)

        return resultados

    def procesar_mensaje(self, mensaje: str) -> Dict:
        gastos = self.procesar_mensaje_multiple(mensaje)
        if gastos:
            return gastos[0]
        return {
            'tipo': 'gasto',
            'categoria': '',
            'concepto': '',
            'monto': 0.0,
            'fecha': datetime.now().strftime('%Y-%m-%d'),
            'metodo_pago': 'Efectivo',
            'notas': mensaje,
            'es_gasto_fijo': False,
        }

    def generar_respuesta(self, resultado: Dict) -> str:
        if resultado['es_gasto_fijo']:
            return (
                f'Gasto fijo registrado:\n'
                f'Concepto: {resultado["concepto"]}\n'
                f'Monto: ${resultado["monto"]:,.0f} COP\n'
                f'Categoria: {resultado["categoria"]}\n'
                f'Fecha: {resultado["fecha"]}'
            )

        return (
            f'Gasto registrado:\n'
            f'Concepto: {resultado["concepto"]}\n'
            f'Monto: ${resultado["monto"]:,.0f} COP\n'
            f'Categoria: {resultado["categoria"]}\n'
            f'Metodo de pago: {resultado["metodo_pago"]}\n'
            f'Fecha: {resultado["fecha"]}'
        )

    def generar_respuesta_multiple(self, resultados: List[Dict]) -> str:
        total = sum(r.get('monto', 0) for r in resultados)
        lineas = [f'Se registraron {len(resultados)} gastos por ${total:,.0f} COP:']
        for r in resultados[:8]:
            lineas.append(f'- {r["concepto"]}: ${r["monto"]:,.0f} ({r["categoria"]})')
        if len(resultados) > 8:
            lineas.append(f'- ... y {len(resultados) - 8} mas')
        return '\n'.join(lineas)

    def obtener_ayuda(self) -> str:
        return (
            '*COMANDOS DISPONIBLES:*\n\n'
            '*Para registrar gastos (uno o varios):*\n'
            '- Gaste 25000 en transporte\n'
            '- Almuerzo 18000 y uber 12000\n'
            '- Pague 45000 de netflix, 12000 de taxi y 9000 de cafe\n'
            '- Mercado 85000; farmacia 23000; gasolina 40000\n\n'
            '*Para consultar:*\n'
            '- saldo: Ver sueldo mensual\n'
            '- resumen: Ver resumen del mes\n'
            '- gastos: Ver lista de gastos del mes\n\n'
            '*Para actualizar datos:*\n'
            '- sueldo [monto]: Cambiar sueldo mensual\n'
            '- gasto [concepto] [monto]: Actualizar gasto fijo'
        )

    def es_comando(self, mensaje: str) -> bool:
        comandos = ['saldo', 'resumen', 'gastos', 'ayuda', 'sueldo', 'gasto', 'eliminar']
        return any(mensaje.lower().startswith(cmd) for cmd in comandos)

    def procesar_comando(self, mensaje: str) -> Dict:
        txt = mensaje.lower().strip()

        if txt.startswith('saldo'):
            return {'tipo': 'consulta', 'accion': 'saldo'}
        if txt.startswith('resumen'):
            return {'tipo': 'consulta', 'accion': 'resumen'}
        if txt.startswith('gastos'):
            return {'tipo': 'consulta', 'accion': 'lista_gastos'}
        if txt.startswith('ayuda') or txt.startswith('help'):
            return {'tipo': 'ayuda', 'mensaje': self.obtener_ayuda()}
        if txt.startswith('sueldo'):
            partes = mensaje.split()
            if len(partes) <= 1:
                return {'tipo': 'error', 'mensaje': 'Debes especificar el monto. Ejemplo: sueldo 5000000'}
            try:
                nuevo_sueldo = self.extraer_monto(partes[1])
                return {'tipo': 'configuracion', 'accion': 'cambiar_sueldo', 'valor': nuevo_sueldo}
            except Exception:
                return {'tipo': 'error', 'mensaje': 'Formato incorrecto. Usa: sueldo 5000000'}

        if txt.startswith('gasto'):
            partes = mensaje.split(maxsplit=2)
            if len(partes) < 3:
                return {'tipo': 'error', 'mensaje': 'Formato incorrecto. Usa: gasto [concepto] [monto]'}
            try:
                monto = self.extraer_monto(partes[2])
                return {'tipo': 'configuracion', 'accion': 'actualizar_gasto', 'concepto': partes[1], 'valor': monto}
            except Exception:
                return {'tipo': 'error', 'mensaje': 'Formato incorrecto. Usa: gasto netflix 30000'}

        return {
            'tipo': 'desconocido',
            'mensaje': 'Comando no reconocido. Escribe "ayuda" para ver los comandos disponibles.',
        }


class GestorExcel:
    def __init__(self, config_path='config/configuracion.json'):
        self.config_path = config_path
        self.temp_dir = os.path.join(tempfile.gettempdir(), 'control_gastos')
        os.makedirs(self.temp_dir, exist_ok=True)
        self.archivo_temp = os.path.join(self.temp_dir, 'ControlDeGastos.xlsx')
        self.meses = [
            'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
        ]

    def _nombre_hoja_actual(self):
        ahora = datetime.now()
        mes_actual = self.meses[ahora.month - 1]
        return f'{mes_actual} {ahora.year}', mes_actual, ahora.year

    def descargar_excel_drive(self) -> str:
        try:
            try:
                from google_drive_v2 import GoogleDriveManager
            except ModuleNotFoundError:
                from src.google_drive_v2 import GoogleDriveManager

            drive = GoogleDriveManager(self.config_path)
            if not drive.autenticar():
                return None
            return drive.descargar_excel_drive(self.archivo_temp)
        except Exception as e:
            print(f'Error descargando de Drive: {e}')
            return None

    def subir_excel_drive(self, ruta_local: str) -> bool:
        try:
            try:
                from google_drive_v2 import GoogleDriveManager
            except ModuleNotFoundError:
                from src.google_drive_v2 import GoogleDriveManager

            drive = GoogleDriveManager(self.config_path)
            if not drive.autenticar():
                return False

            file_id = drive.subir_excel_drive(ruta_local, actualizar=True)
            return file_id is not None
        except Exception as e:
            print(f'Error subiendo a Drive: {e}')
            return False

    def agregar_gastos(self, gastos: List[Dict]) -> bool:
        from openpyxl import load_workbook
        try:
            from excel_mensual import GeneradorExcelMensual
        except ModuleNotFoundError:
            from src.excel_mensual import GeneradorExcelMensual

        if not gastos:
            return False

        try:
            print(f'Procesando {len(gastos)} gasto(s)...')
            print('Descargando Excel desde Drive...')
            ruta_excel = self.descargar_excel_drive()

            generador = GeneradorExcelMensual(self.config_path)
            if not ruta_excel or not os.path.exists(ruta_excel):
                print('Creando nuevo Excel...')
                wb = generador.crear_excel_nuevo()
                ruta_excel = generador.guardar_excel_temporal(wb)
                print(f'Excel creado en: {ruta_excel}')
            else:
                print(f'Cargando Excel existente: {ruta_excel}')
                wb = load_workbook(ruta_excel)

            hoja_actual, mes_actual, anio_actual = self._nombre_hoja_actual()
            print(f'Mes actual: {mes_actual} {anio_actual}')
            print(f'Hojas disponibles: {wb.sheetnames}')

            if hoja_actual not in wb.sheetnames and mes_actual not in wb.sheetnames:
                print(f'Creando hoja para {hoja_actual}...')
            elif hoja_actual not in wb.sheetnames and mes_actual in wb.sheetnames:
                print(f'Migrando hoja legacy "{mes_actual}" a "{hoja_actual}"...')
            else:
                print(f'Actualizando estructura de hoja {hoja_actual} sin perder registros...')

            ws = generador.crear_o_actualizar_hoja_mes(wb, mes_actual, anio_actual)
            for gasto in gastos:
                generador.agregar_gasto_a_hoja(ws, gasto)

            print('Guardando Excel local...')
            wb.save(ruta_excel)
            print(f'Excel guardado en: {ruta_excel}')

            print('Sincronizando con Google Drive...')
            exito_sync = self.subir_excel_drive(ruta_excel)
            if exito_sync:
                print('Sincronizado con Drive exitosamente')
            else:
                print('Error sincronizando con Drive, pero los gastos se guardaron localmente')

            return True
        except Exception as e:
            print(f'ERROR al agregar gastos: {e}')
            import traceback
            traceback.print_exc()
            return False

    def agregar_gasto(self, datos_gasto: Dict) -> bool:
        return self.agregar_gastos([datos_gasto])

    def obtener_resumen(self) -> Dict:
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)

            return {
                'ingresos': config.get('sueldo', {}).get('valor_fijo', 0),
                'gastos_fijos': sum(g.get('valor', 0) for g in config.get('gastos_fijos', {}).values()),
                'saldo_bancario': config.get('saldo_bancario', {}).get('valor_actual', 0),
            }
        except Exception as e:
            return {'error': str(e)}

    def sincronizar_con_drive(self) -> bool:
        try:
            try:
                from google_drive_v2 import sincronizar_con_drive
            except ModuleNotFoundError:
                from src.google_drive_v2 import sincronizar_con_drive
            return sincronizar_con_drive(self.config_path)
        except Exception as e:
            print(f'Error en sincronizacion: {e}')
            return False


class BotWhatsApp:
    def __init__(self):
        self.procesador = ProcesadorMensajes()
        self.gestor_excel = GestorExcel()

    def procesar_entrada(self, mensaje: str, numero_remitente: str = None) -> str:
        if self.procesador.es_comando(mensaje):
            resultado = self.procesador.procesar_comando(mensaje)
            if resultado['tipo'] == 'ayuda':
                return resultado['mensaje']
            if resultado['tipo'] == 'consulta':
                return self._manejar_consulta(resultado['accion'])
            if resultado['tipo'] == 'configuracion':
                return self._manejar_configuracion(resultado)
            if resultado['tipo'] == 'error':
                return resultado['mensaje']
            return resultado.get('mensaje', 'Comando no reconocido')

        gastos = self.procesador.procesar_mensaje_multiple(mensaje)
        if not gastos:
            return (
                'No pude identificar montos en tu mensaje.\n'
                'Ejemplos: "almuerzo 18000 y uber 12000" o "mercado 85000; farmacia 23000".'
            )

        exito = self.gestor_excel.agregar_gastos(gastos)
        if not exito:
            return 'Error al registrar los gastos. Verifica la conexion y el archivo Excel.'

        if len(gastos) == 1:
            return self.procesador.generar_respuesta(gastos[0])
        return self.procesador.generar_respuesta_multiple(gastos)

    def _manejar_consulta(self, accion: str) -> str:
        if accion == 'saldo':
            resumen = self.gestor_excel.obtener_resumen()
            if 'error' in resumen:
                return f'Error: {resumen["error"]}'
            return f'Tu sueldo mensual es: ${resumen["ingresos"]:,.0f} COP'
        if accion == 'resumen':
            return 'Consulta el resumen en tu hoja mensual del Excel.'
        if accion == 'lista_gastos':
            return 'Consulta la tabla "GASTOS VARIABLES DEL MES" en la hoja del mes.'
        return 'Consulta no reconocida'

    def _manejar_configuracion(self, datos: Dict) -> str:
        if datos['accion'] == 'cambiar_sueldo':
            try:
                from automatizador import AutomatizadorGastos
            except ModuleNotFoundError:
                from src.automatizador import AutomatizadorGastos
            auto = AutomatizadorGastos()
            auto.cambiar_sueldo(datos['valor'])
            return f'Sueldo actualizado a ${datos["valor"]:,.0f} COP'
        if datos['accion'] == 'actualizar_gasto':
            try:
                from automatizador import AutomatizadorGastos
            except ModuleNotFoundError:
                from src.automatizador import AutomatizadorGastos
            auto = AutomatizadorGastos()
            auto.actualizar_gasto_fijo(datos['concepto'], datos['valor'])
            return f'Gasto fijo "{datos["concepto"]}" actualizado a ${datos["valor"]:,.0f} COP'
        return 'Configuracion no reconocida'


if __name__ == '__main__':
    bot = BotWhatsApp()

    print('=== BOT DE CONTROL DE GASTOS ===\n')
    print('Escribe tus gastos o comandos (escribe "ayuda" para ver opciones)')
    print('Escribe "salir" para terminar\n')

    while True:
        mensaje = input('Tu: ')
        if mensaje.lower() == 'salir':
            print('Hasta luego')
            break
        respuesta = bot.procesar_entrada(mensaje)
        print(f'\nBot: {respuesta}\n')
