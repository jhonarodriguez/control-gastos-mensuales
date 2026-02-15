import json
import os
import tempfile
import time
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class GeneradorExcelMensual:
    """
    Generador de Excel mensual:
    - 1 hoja por mes (formato "Mes Año")
    - vista tipo tablero con resumen, fijos, deudas y variables
    - registro de gastos del bot solo en variables del mes
    """

    MESES = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]

    FILA_FIJOS_DATA_INICIO = 4
    FILA_FIJOS_DATA_FIN = 23
    FILA_FIJOS_TOTAL = 24

    FILA_DEUDAS_DATA_INICIO = 4
    FILA_DEUDAS_DATA_FIN = 23
    FILA_DEUDAS_TOTAL = 24

    FILA_VARIABLES_DATA_INICIO = 4
    FILA_VARIABLES_DATA_FIN = 28
    FILA_VARIABLES_TOTAL = 29

    DEFAULT_RETIRO_EFECTIVO_ITEMS = ['gasto:arriendo']
    DEFAULT_MOVII_ITEMS = [
        'gasto:netflix',
        'gasto:youtube_premium',
        'gasto:google_drive',
        'gasto:mercadolibre',
        'gasto:hbo_max',
        'gasto:pago_app_fitia',
        'gasto:sub_facebook_don_j',
    ]

    def __init__(self, config_path='config/configuracion.json'):
        self.config_path = config_path
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = json.load(f)

        self.colores = {
            'titulo': '1F4E78',
            'resumen': '2E7D32',
            'fijos_header': 'F4B400',
            'deudas_header': 'DB4437',
            'variables_header': 'FB8C00',
            'subheader': '5C6BC0',
            'total': '2E7D32',
            'ingreso': 'C8E6C9',
            'saldo': 'B3E5FC',
            'alerta': 'FFCDD2',
            'notas': 'FFF59D',
            'blanco': 'FFFFFF',
        }

        self.borde = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

    def _nombre_hoja_mes(self, mes_nombre, anio):
        return f'{mes_nombre} {anio}'

    def _colorear(self, ws, ref, color, bold=False, font_color='000000', align='left'):
        cell = ws[ref]
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.font = Font(bold=bold, color=font_color)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = self.borde

    def _normalizar_numero(self, valor):
        if isinstance(valor, (int, float)):
            return float(valor)
        try:
            if valor is None or valor == '':
                return 0.0
            txt = str(valor).replace('$', '').replace(',', '').strip()
            return float(txt)
        except Exception:
            return 0.0

    def _obtener_saldo_inicio_mes(self, mes_nombre, anio):
        saldo_real = self._normalizar_numero(
            self.config.get('saldo_bancario', {}).get('valor_actual', 0)
        )
        historial = self.config.get('historial_saldos', {})
        saldos_mensuales = historial.get('saldos_mensuales', {})

        try:
            mes_idx = self.MESES.index(mes_nombre) + 1
        except ValueError:
            mes_idx = datetime.now().month

        # Prioridad 1: saldo inicial configurado manualmente para el mes actual
        llave_actual = f'{anio:04d}-{mes_idx:02d}'
        registro_actual = self._obtener_registro_mes(historial, llave_actual)
        saldo_inicio_actual = self._normalizar_numero(
            registro_actual.get('saldo_inicial', 0)
        )
        if saldo_inicio_actual > 0:
            return saldo_inicio_actual

        if mes_idx == 1:
            anio_prev = anio - 1
            mes_prev = 12
        else:
            anio_prev = anio
            mes_prev = mes_idx - 1

        llave_prev = f'{anio_prev:04d}-{mes_prev:02d}'
        registro_prev = self._obtener_registro_mes(historial, llave_prev)
        saldo_prev = self._normalizar_numero(
            registro_prev.get('saldo_final', 0)
        )

        if saldo_prev > 0:
            return saldo_prev

        saldo_mes_anterior = self._normalizar_numero(historial.get('saldo_mes_anterior', 0))
        if saldo_mes_anterior > 0:
            return saldo_mes_anterior

        return saldo_real

    def _obtener_registro_mes(self, historial, llave_mes):
        if not isinstance(historial, dict):
            return {}
        saldos_mensuales = historial.get('saldos_mensuales', {})
        if isinstance(saldos_mensuales, dict):
            registro = saldos_mensuales.get(llave_mes, {})
            if isinstance(registro, dict):
                return registro
        # Compatibilidad con estructura legacy: historial_saldos["YYYY-MM"]
        legacy = historial.get(llave_mes, {})
        return legacy if isinstance(legacy, dict) else {}

    def _obtener_ingresos_extra_mes(self, mes_nombre, anio):
        historial = self.config.get('historial_saldos', {})

        try:
            mes_idx = self.MESES.index(mes_nombre) + 1
        except ValueError:
            mes_idx = datetime.now().month

        llave_actual = f'{anio:04d}-{mes_idx:02d}'
        registro_mes = self._obtener_registro_mes(historial, llave_actual)
        ingresos_raw = registro_mes.get('ingresos_extra', [])
        if not isinstance(ingresos_raw, list):
            return [], 0.0

        ingresos = []
        total = 0.0
        for item in ingresos_raw:
            if not isinstance(item, dict):
                continue
            concepto = str(item.get('concepto', 'Ingreso extra')).strip() or 'Ingreso extra'
            valor = self._normalizar_numero(item.get('valor', 0))
            fecha_registro = str(item.get('fecha_registro', '') or '')
            if valor <= 0:
                continue
            ingresos.append({'concepto': concepto, 'valor': valor, 'fecha_registro': fecha_registro})
            total += valor

        return ingresos, total

    def _obtener_compromisos_por_id(self):
        compromisos = {}

        for key, datos in (self.config.get('gastos_fijos') or {}).items():
            item_id = f'gasto:{key}'
            compromisos[item_id] = {
                'id': item_id,
                'tipo': 'Gasto fijo',
                'key': key,
                'concepto': str(key).replace('_', ' ').title(),
                'valor': self._normalizar_numero((datos or {}).get('valor', 0)),
            }

        for key, datos in (self.config.get('deudas_fijas') or {}).items():
            item_id = f'deuda:{key}'
            compromisos[item_id] = {
                'id': item_id,
                'tipo': 'Deuda',
                'key': key,
                'concepto': str(key).replace('_', ' ').title(),
                'valor': self._normalizar_numero((datos or {}).get('valor', 0)),
            }

        return compromisos

    def _normalizar_item_flujo(self, raw_item, compromisos):
        if not isinstance(raw_item, str):
            return None
        item = raw_item.strip()
        if not item:
            return None
        if item in compromisos:
            return item

        # Compatibilidad con formato legacy solo por nombre de clave.
        legacy_key = item.lower().replace(' ', '_')
        gasto_id = f'gasto:{legacy_key}'
        deuda_id = f'deuda:{legacy_key}'
        if gasto_id in compromisos:
            return gasto_id
        if deuda_id in compromisos:
            return deuda_id
        return None

    def _obtener_resumen_flujo(self, clave_items, defaults):
        compromisos = self._obtener_compromisos_por_id()
        flujos = self.config.get('flujos_efectivo', {}) or {}
        seleccion_raw = flujos.get(clave_items, defaults)
        if not isinstance(seleccion_raw, list):
            seleccion_raw = defaults

        seleccion = []
        vistos = set()
        for raw_item in seleccion_raw:
            item_id = self._normalizar_item_flujo(raw_item, compromisos)
            if not item_id or item_id in vistos:
                continue
            seleccion.append(item_id)
            vistos.add(item_id)

        detalle = []
        total = 0.0
        for item_id in seleccion:
            item = compromisos.get(item_id)
            if not item:
                continue
            valor = self._normalizar_numero(item.get('valor', 0))
            if valor <= 0:
                continue
            total += valor
            detalle.append({
                'id': item_id,
                'concepto': item.get('concepto', item_id),
                'valor': valor,
                'tipo': item.get('tipo', ''),
            })

        return detalle, total

    def _resumir_detalle_flujo(self, detalle, max_items=3):
        if not detalle:
            return 'Sin elementos seleccionados'
        items = [f'{x["concepto"]}: ${x["valor"]:,.0f}' for x in detalle[:max_items]]
        resumen = ' | '.join(items)
        if len(detalle) > max_items:
            resumen += f' | +{len(detalle) - max_items} mas'
        return resumen

    def _extraer_registros_existentes(self, ws):
        variables = []

        # Extraer variables del tablero (K:M)
        for fila in range(self.FILA_VARIABLES_DATA_INICIO, self.FILA_VARIABLES_DATA_FIN + 1):
            monto = ws[f'K{fila}'].value
            concepto = ws[f'L{fila}'].value
            meta = ws[f'M{fila}'].value

            if concepto and str(concepto).strip():
                monto_num = self._normalizar_numero(monto)
                if monto_num > 0:
                    variables.append((monto_num, str(concepto), '' if meta is None else str(meta)))

        # Compatibilidad: si no hay variables, intentar reconstruir desde una tabla
        # legacy "DETALLE DE GASTOS (BOT)" en columnas A:F.
        if not variables:
            fila_titulo = None
            for fila in range(1, ws.max_row + 1):
                valor = ws[f'A{fila}'].value
                if valor and 'DETALLE DE GASTOS' in str(valor).upper():
                    fila_titulo = fila
                    break

            if fila_titulo:
                fila = fila_titulo + 2
                while fila <= ws.max_row:
                    fecha = ws[f'A{fila}'].value
                    concepto = ws[f'B{fila}'].value
                    categoria = ws[f'C{fila}'].value
                    monto = ws[f'D{fila}'].value
                    metodo = ws[f'E{fila}'].value
                    notas = ws[f'F{fila}'].value

                    if not any([fecha, concepto, categoria, monto, metodo, notas]):
                        break

                    if fecha and 'EJEMPLO' in str(fecha).upper():
                        fila += 1
                        continue

                    monto_num = self._normalizar_numero(monto)
                    if monto_num > 0:
                        fecha_txt = '' if fecha is None else str(fecha)
                        categoria_txt = '' if categoria is None else str(categoria)
                        variables.append((
                            monto_num,
                            '' if concepto is None else str(concepto),
                            f'{categoria_txt} | {fecha_txt}'.strip(' |'),
                        ))
                    fila += 1

        return variables

    def _limpiar_hoja(self, ws):
        # Limpieza total de contenido y estilos para reconstrucción consistente
        if ws.merged_cells.ranges:
            for rng in list(ws.merged_cells.ranges):
                try:
                    ws.unmerge_cells(str(rng))
                except Exception:
                    pass
        ws.delete_rows(1, ws.max_row + 1)

    def _escribir_fijos(self, ws):
        gastos_fijos = self.config.get('gastos_fijos', {})
        fila = self.FILA_FIJOS_DATA_INICIO
        for nombre, datos in gastos_fijos.items():
            if fila > self.FILA_FIJOS_DATA_FIN:
                break
            monto = self._normalizar_numero(datos.get('valor', 0))
            categoria = datos.get('categoria', 'Sin categoría')
            fecha = ''
            if 'dia_cargo' in datos:
                fecha = f"Día {datos.get('dia_cargo')}"
            elif 'frecuencia' in datos:
                fecha = str(datos.get('frecuencia', '')).title()

            ws[f'E{fila}'] = monto
            ws[f'F{fila}'] = nombre.replace('_', ' ').title()
            ws[f'G{fila}'] = f'{categoria} | {fecha}'.strip(' |')
            ws[f'E{fila}'].number_format = '$#,##0'

            self._colorear(ws, f'E{fila}', 'FCE8B2')
            self._colorear(ws, f'F{fila}', 'FCE8B2')
            self._colorear(ws, f'G{fila}', self.colores['notas'])
            fila += 1

        # Mantener celdas de tabla con bordes para edición manual
        for f in range(fila, self.FILA_FIJOS_DATA_FIN + 1):
            self._colorear(ws, f'E{f}', 'FCE8B2')
            self._colorear(ws, f'F{f}', 'FCE8B2')
            self._colorear(ws, f'G{f}', self.colores['notas'])

    def _escribir_deudas(self, ws):
        deudas = self.config.get('deudas_fijas', {})
        fila = self.FILA_DEUDAS_DATA_INICIO
        for nombre, datos in deudas.items():
            if fila > self.FILA_DEUDAS_DATA_FIN:
                break
            monto = self._normalizar_numero(datos.get('valor', 0))
            fecha = ''
            if 'dia_cargo' in datos:
                fecha = f"Día {datos.get('dia_cargo')}"
            elif 'frecuencia' in datos:
                fecha = str(datos.get('frecuencia', '')).title()
            detalle = datos.get('detalle', '')

            ws[f'H{fila}'] = monto
            ws[f'I{fila}'] = nombre.replace('_', ' ').title()
            ws[f'J{fila}'] = f'{fecha} {detalle}'.strip()
            ws[f'H{fila}'].number_format = '$#,##0'

            self._colorear(ws, f'H{fila}', 'F8D7DA')
            self._colorear(ws, f'I{fila}', 'F8D7DA')
            self._colorear(ws, f'J{fila}', self.colores['notas'])
            fila += 1

        for f in range(fila, self.FILA_DEUDAS_DATA_FIN + 1):
            self._colorear(ws, f'H{f}', 'F8D7DA')
            self._colorear(ws, f'I{f}', 'F8D7DA')
            self._colorear(ws, f'J{f}', self.colores['notas'])

    def _construir_layout_base(self, ws, mes_nombre, anio):
        sueldo = self._normalizar_numero(self.config.get('sueldo', {}).get('valor_fijo', 0))
        saldo_real = self._normalizar_numero(self.config.get('saldo_bancario', {}).get('valor_actual', 0))
        saldo_inicio = self._obtener_saldo_inicio_mes(mes_nombre, anio)
        ingresos_extra_detalle, ingresos_extra_total = self._obtener_ingresos_extra_mes(mes_nombre, anio)
        retiro_detalle, retiro_total = self._obtener_resumen_flujo(
            'retiro_efectivo_items',
            self.DEFAULT_RETIRO_EFECTIVO_ITEMS,
        )
        movii_detalle, movii_total = self._obtener_resumen_flujo(
            'movii_items',
            self.DEFAULT_MOVII_ITEMS,
        )
        presupuesto_variables = self._normalizar_numero(self.config.get('presupuesto_variables', 0))

        # Título principal
        ws['A1'] = f'CONTROL DE GASTOS - {mes_nombre.upper()} {anio}'
        ws.merge_cells('A1:M1')
        self._colorear(ws, 'A1', self.colores['titulo'], bold=True, font_color='FFFFFF', align='center')
        ws.row_dimensions[1].height = 32

        # Encabezados de bloques
        ws.merge_cells('A2:C2')
        ws['A2'] = 'RESUMEN DEL MES'
        self._colorear(ws, 'A2', self.colores['resumen'], bold=True, font_color='FFFFFF', align='center')

        ws.merge_cells('E2:G2')
        ws['E2'] = 'FIJOS CONFIGURADOS'
        self._colorear(ws, 'E2', self.colores['fijos_header'], bold=True, align='center')

        ws.merge_cells('H2:J2')
        ws['H2'] = 'DEUDAS MENSUALES'
        self._colorear(ws, 'H2', self.colores['deudas_header'], bold=True, font_color='FFFFFF', align='center')

        ws.merge_cells('K2:M2')
        ws['K2'] = 'GASTOS VARIABLES DEL MES'
        self._colorear(ws, 'K2', self.colores['variables_header'], bold=True, align='center')

        # Subheaders de tablas laterales
        for ref, txt in [('E3', 'Monto'), ('F3', 'Concepto'), ('G3', 'Categoría / Fecha')]:
            ws[ref] = txt
            self._colorear(ws, ref, self.colores['subheader'], bold=True, font_color='FFFFFF', align='center')
        for ref, txt in [('H3', 'Monto'), ('I3', 'Concepto'), ('J3', 'Fecha / Nota')]:
            ws[ref] = txt
            self._colorear(ws, ref, self.colores['subheader'], bold=True, font_color='FFFFFF', align='center')
        for ref, txt in [('K3', 'Monto'), ('L3', 'Concepto'), ('M3', 'Categoría / Fecha')]:
            ws[ref] = txt
            self._colorear(ws, ref, self.colores['subheader'], bold=True, font_color='FFFFFF', align='center')

        # Resumen mensual (A:B)
        resumen_labels = [
            (3, 'Mes de trabajo', f'{mes_nombre} {anio}', self.colores['saldo']),
            (4, 'Sueldo mensual (base)', sueldo, self.colores['ingreso']),
            (5, 'Ingresos extra del mes', ingresos_extra_total, self.colores['ingreso']),
            (6, 'Ingreso total del mes', '=B4+B5', self.colores['ingreso']),
            (7, 'Saldo inicio de mes', saldo_inicio, self.colores['saldo']),
            (8, 'Total gastos fijos', f'=E{self.FILA_FIJOS_TOTAL}', self.colores['alerta']),
            (9, 'Total deudas', f'=H{self.FILA_DEUDAS_TOTAL}', self.colores['alerta']),
            (10, 'Total variables', f'=K{self.FILA_VARIABLES_TOTAL}', self.colores['variables_header']),
            (11, 'TOTAL GASTOS MES', '=B8+B9+B10', 'FFB74D'),
            (12, 'Saldo proyectado (inicio + ingresos - gastos)', '=B7+B6-B11', self.colores['total']),
            (13, 'Saldo REAL en banco', saldo_real, 'FFD54F'),
            (14, 'Diferencia (real - proyectado)', '=B13-B12', self.colores['alerta']),
            (15, 'Presupuesto variables', presupuesto_variables, self.colores['saldo']),
            (16, '% gasto sobre ingreso total', '=IF(B6=0,0,B11/B6)', self.colores['saldo']),
            (17, '% variables sobre ingreso total', '=IF(B6=0,0,B10/B6)', self.colores['saldo']),
            (18, 'Estado presupuesto variables', '=B15-B10', self.colores['saldo']),
            (19, 'Retiro en efectivo (configurado)', retiro_total, self.colores['saldo']),
            (20, 'Recarga MOVII (configurada)', movii_total, self.colores['saldo']),
            (21, 'Total flujos programados', '=B19+B20', self.colores['saldo']),
        ]

        for fila, label, value, color in resumen_labels:
            ws[f'A{fila}'] = label
            ws[f'B{fila}'] = value
            self._colorear(ws, f'A{fila}', color, bold=('TOTAL' in label or 'Diferencia' in label))
            self._colorear(ws, f'B{fila}', color, bold=('TOTAL' in label or 'Diferencia' in label))

            if fila in (4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 18, 19, 20, 21):
                ws[f'B{fila}'].number_format = '$#,##0'
            if fila in (16, 17):
                ws[f'B{fila}'].number_format = '0.00%'

        # Columna C para notas rápidas
        ws['C3'] = 'Notas del mes'
        self._colorear(ws, 'C3', self.colores['subheader'], bold=True, font_color='FFFFFF', align='center')
        for fila in range(4, 22):
            self._colorear(ws, f'C{fila}', self.colores['notas'])
        if ingresos_extra_detalle:
            detalles_txt = ' | '.join([f'{x["concepto"]}: ${x["valor"]:,.0f}' for x in ingresos_extra_detalle[:3]])
            if len(ingresos_extra_detalle) > 3:
                detalles_txt += f' | +{len(ingresos_extra_detalle) - 3} más'
            ws['C5'] = detalles_txt
        ws['C19'] = self._resumir_detalle_flujo(retiro_detalle)
        ws['C20'] = self._resumir_detalle_flujo(movii_detalle)

        # Totales de tablas laterales
        ws[f'E{self.FILA_FIJOS_TOTAL}'] = f'=SUM(E{self.FILA_FIJOS_DATA_INICIO}:E{self.FILA_FIJOS_DATA_FIN})'
        ws[f'F{self.FILA_FIJOS_TOTAL}'] = 'TOTAL FIJOS'
        ws[f'H{self.FILA_DEUDAS_TOTAL}'] = f'=SUM(H{self.FILA_DEUDAS_DATA_INICIO}:H{self.FILA_DEUDAS_DATA_FIN})'
        ws[f'I{self.FILA_DEUDAS_TOTAL}'] = 'TOTAL DEUDAS'
        ws[f'K{self.FILA_VARIABLES_TOTAL}'] = f'=SUM(K{self.FILA_VARIABLES_DATA_INICIO}:K{self.FILA_VARIABLES_DATA_FIN})'
        ws[f'L{self.FILA_VARIABLES_TOTAL}'] = 'TOTAL VARIABLES'

        for ref in [f'E{self.FILA_FIJOS_TOTAL}', f'F{self.FILA_FIJOS_TOTAL}', f'G{self.FILA_FIJOS_TOTAL}']:
            self._colorear(ws, ref, self.colores['total'], bold=True, font_color='FFFFFF')
        for ref in [f'H{self.FILA_DEUDAS_TOTAL}', f'I{self.FILA_DEUDAS_TOTAL}', f'J{self.FILA_DEUDAS_TOTAL}']:
            self._colorear(ws, ref, self.colores['total'], bold=True, font_color='FFFFFF')
        for ref in [f'K{self.FILA_VARIABLES_TOTAL}', f'L{self.FILA_VARIABLES_TOTAL}', f'M{self.FILA_VARIABLES_TOTAL}']:
            self._colorear(ws, ref, self.colores['total'], bold=True, font_color='FFFFFF')

        ws[f'E{self.FILA_FIJOS_TOTAL}'].number_format = '$#,##0'
        ws[f'H{self.FILA_DEUDAS_TOTAL}'].number_format = '$#,##0'
        ws[f'K{self.FILA_VARIABLES_TOTAL}'].number_format = '$#,##0'

        # Indicadores extra
        ws.merge_cells('A22:C22')
        ws['A22'] = 'INDICADORES Y CONTROL'
        self._colorear(ws, 'A22', self.colores['titulo'], bold=True, font_color='FFFFFF', align='center')

        indicadores = [
            (23, 'Ahorro del ingreso total (ingresos - gastos)', '=B6-B11'),
            (24, 'Patrimonio estimado fin mes', '=B12'),
            (25, 'Estado banco vs sistema', '=IF(B14=0,"CUADRADO",IF(B14>0,"SOBRANTE","FALTANTE"))'),
            (26, 'Sugerencia ahorro (20% ingreso total)', '=B6*0.2'),
            (27, 'Desviación vs sugerencia', '=B23-B26'),
        ]
        for fila, label, formula in indicadores:
            ws[f'A{fila}'] = label
            ws[f'B{fila}'] = formula
            self._colorear(ws, f'A{fila}', self.colores['saldo'])
            self._colorear(ws, f'B{fila}', self.colores['saldo'], bold=(fila in (23, 25)))
            if fila in (23, 24, 26, 27):
                ws[f'B{fila}'].number_format = '$#,##0'

        for fila in range(23, 28):
            self._colorear(ws, f'C{fila}', self.colores['notas'])

        # Bloque visible para ingresos extra del mes
        ws.merge_cells('A29:C29')
        ws['A29'] = 'INGRESOS EXTRA DEL MES (DETALLE)'
        self._colorear(ws, 'A29', self.colores['titulo'], bold=True, font_color='FFFFFF', align='center')

        ws['A30'] = 'Concepto'
        ws['B30'] = 'Valor'
        ws['C30'] = 'Fecha registro'
        self._colorear(ws, 'A30', self.colores['subheader'], bold=True, font_color='FFFFFF', align='center')
        self._colorear(ws, 'B30', self.colores['subheader'], bold=True, font_color='FFFFFF', align='center')
        self._colorear(ws, 'C30', self.colores['subheader'], bold=True, font_color='FFFFFF', align='center')

        fila_ing = 31
        max_fila_ing = 35
        if ingresos_extra_detalle:
            for ing in ingresos_extra_detalle:
                if fila_ing > max_fila_ing:
                    break
                ws[f'A{fila_ing}'] = ing.get('concepto', 'Ingreso extra')
                ws[f'B{fila_ing}'] = self._normalizar_numero(ing.get('valor', 0))
                ws[f'C{fila_ing}'] = ing.get('fecha_registro', '')
                ws[f'B{fila_ing}'].number_format = '$#,##0'
                self._colorear(ws, f'A{fila_ing}', self.colores['ingreso'])
                self._colorear(ws, f'B{fila_ing}', self.colores['ingreso'])
                self._colorear(ws, f'C{fila_ing}', self.colores['notas'])
                fila_ing += 1
        else:
            ws['A31'] = 'Sin ingresos extra registrados'
            ws.merge_cells('A31:C31')
            self._colorear(ws, 'A31', self.colores['notas'])
            fila_ing = 32

        for fila in range(fila_ing, max_fila_ing + 1):
            self._colorear(ws, f'A{fila}', self.colores['ingreso'])
            self._colorear(ws, f'B{fila}', self.colores['ingreso'])
            self._colorear(ws, f'C{fila}', self.colores['notas'])

        # Anchos y paneles
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 14
        ws.column_dimensions['L'].width = 28
        ws.column_dimensions['M'].width = 22

        ws.freeze_panes = 'A4'

    def _insertar_registros_preservados(self, ws, variables):
        fila_var = self.FILA_VARIABLES_DATA_INICIO
        for monto, concepto, meta in variables:
            if fila_var > self.FILA_VARIABLES_DATA_FIN:
                break
            ws[f'K{fila_var}'] = self._normalizar_numero(monto)
            ws[f'L{fila_var}'] = concepto
            ws[f'M{fila_var}'] = meta
            ws[f'K{fila_var}'].number_format = '$#,##0'
            self._colorear(ws, f'K{fila_var}', 'FFE0B2')
            self._colorear(ws, f'L{fila_var}', 'FFE0B2')
            self._colorear(ws, f'M{fila_var}', self.colores['notas'])
            fila_var += 1

        for fila in range(fila_var, self.FILA_VARIABLES_DATA_FIN + 1):
            self._colorear(ws, f'K{fila}', 'FFE0B2')
            self._colorear(ws, f'L{fila}', 'FFE0B2')
            self._colorear(ws, f'M{fila}', self.colores['notas'])

    def crear_o_actualizar_hoja_mes(self, wb, mes_nombre, anio):
        """
        Crea o actualiza (sin perder gastos) la hoja mensual.
        """
        hoja_objetivo = self._nombre_hoja_mes(mes_nombre, anio)
        hoja_legacy = mes_nombre

        variables = []

        if hoja_objetivo in wb.sheetnames:
            ws = wb[hoja_objetivo]
            variables = self._extraer_registros_existentes(ws)
            self._limpiar_hoja(ws)
        elif hoja_legacy in wb.sheetnames:
            ws = wb[hoja_legacy]
            variables = self._extraer_registros_existentes(ws)
            ws.title = hoja_objetivo
            self._limpiar_hoja(ws)
        else:
            ws = wb.create_sheet(hoja_objetivo)

        self._construir_layout_base(ws, mes_nombre, anio)
        self._escribir_fijos(ws)
        self._escribir_deudas(ws)
        self._insertar_registros_preservados(ws, variables)
        return ws

    def _buscar_siguiente_fila_libre(self, ws, col, fila_inicio, fila_fin):
        for fila in range(fila_inicio, fila_fin + 1):
            valor = ws[f'{col}{fila}'].value
            if valor in (None, ''):
                return fila
            if isinstance(valor, str) and 'EJEMPLO' in valor.upper():
                return fila
        return None

    def agregar_gasto_a_hoja(self, ws, datos_gasto):
        """
        Inserta un gasto solo en la tabla de variables del mes (K:M).
        """
        fila_var = self._buscar_siguiente_fila_libre(
            ws, 'K', self.FILA_VARIABLES_DATA_INICIO, self.FILA_VARIABLES_DATA_FIN
        )
        if fila_var:
            ws[f'K{fila_var}'] = self._normalizar_numero(datos_gasto.get('monto', 0))
            ws[f'L{fila_var}'] = datos_gasto.get('concepto', 'Gasto general')
            ws[f'M{fila_var}'] = f"{datos_gasto.get('categoria', 'Otros')} | {datos_gasto.get('fecha', '')}"
            ws[f'K{fila_var}'].number_format = '$#,##0'
            self._colorear(ws, f'K{fila_var}', 'FFE0B2')
            self._colorear(ws, f'L{fila_var}', 'FFE0B2')
            self._colorear(ws, f'M{fila_var}', self.colores['notas'])

    def crear_excel_nuevo(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ahora = datetime.now()
        mes_actual = self.MESES[ahora.month - 1]
        anio_actual = ahora.year

        self.crear_o_actualizar_hoja_mes(wb, mes_actual, anio_actual)
        return wb

    def guardar_excel_temporal(self, wb, nombre='ControlDeGastos.xlsx'):
        temp_dir = os.path.join(tempfile.gettempdir(), 'control_gastos')
        os.makedirs(temp_dir, exist_ok=True)
        ruta = os.path.join(temp_dir, nombre)

        try:
            wb.save(ruta)
        except PermissionError:
            ruta_alt = os.path.join(temp_dir, f'ControlDeGastos_{int(time.time())}.xlsx')
            wb.save(ruta_alt)
            return ruta_alt

        return ruta


if __name__ == '__main__':
    generador = GeneradorExcelMensual()
    wb = generador.crear_excel_nuevo()
    ruta = generador.guardar_excel_temporal(wb)
    print(f'Excel creado: {ruta}')
