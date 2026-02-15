import os
import json
import time
import io
import re
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import pickle
from datetime import datetime
from pathlib import Path

class GoogleDriveManager:
    """
    Gestor de Google Drive para Control de Gastos
    Maneja un solo archivo Excel con hojas mensuales
    """
    
    SCOPES = [
        'https://www.googleapis.com/auth/drive.file',
        'https://www.googleapis.com/auth/drive.metadata.readonly'
    ]
    
    def __init__(self, config_path='config/configuracion.json'):
        self.config_path = config_path
        self.creds = None
        self.service = None
        self.token_path = 'config/token.pickle'
        self.credentials_path = 'config/credentials.json'
        self.nombre_archivo = 'ControlDeGastos.xlsx'
        self.nombre_carpeta = 'ControlDeGastos'
        
        self._cargar_configuracion()
    
    def _cargar_configuracion(self):
        with open(self.config_path, 'r', encoding='utf-8') as f:
            self.config = json.load(f)
        
        self.archivo_excel_id = self.config['google_drive'].get('archivo_excel_id', '')
        self.carpeta_id = self.config['google_drive'].get('carpeta_backup_id', '')
    
    def _guardar_configuracion(self):
        self.config['google_drive']['archivo_excel_id'] = self.archivo_excel_id
        self.config['google_drive']['carpeta_backup_id'] = self.carpeta_id
        
        with open(self.config_path, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, indent=2, ensure_ascii=False)
    
    def autenticar(self):
        """Autentica con Google Drive"""
        if os.path.exists(self.token_path):
            with open(self.token_path, 'rb') as token:
                self.creds = pickle.load(token)
        
        if not self.creds or not self.creds.valid:
            if self.creds and self.creds.expired and self.creds.refresh_token:
                self.creds.refresh(Request())
            else:
                if not os.path.exists(self.credentials_path):
                    print('ERROR: No se encontrÃ³ el archivo credentials.json')
                    print('Por favor descarga tus credenciales de Google Cloud Console')
                    print('y guÃ¡rdalas en config/credentials.json')
                    return False
                
                flow = InstalledAppFlow.from_client_secrets_file(
                    self.credentials_path, self.SCOPES)
                self.creds = flow.run_local_server(port=0)
            
            with open(self.token_path, 'wb') as token:
                pickle.dump(self.creds, token)
        
        self.service = build('drive', 'v3', credentials=self.creds)
        print('AutenticaciÃ³n exitosa con Google Drive')
        return True
    
    def crear_o_obtener_carpeta(self, nombre=None):
        """Crea o obtiene la carpeta de ControlDeGastos"""
        if nombre is None:
            nombre = self.nombre_carpeta
        
        if not self.service:
            print('Error: No has iniciado sesiÃ³n.')
            return None
        
        # Buscar si la carpeta ya existe
        try:
            query = f"mimeType='application/vnd.google-apps.folder' and name='{nombre}' and trashed=false"
            results = self.service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
            items = results.get('files', [])
            
            if items:
                print(f'Carpeta encontrada: {nombre} (ID: {items[0]["id"]})')
                self.carpeta_id = items[0]['id']
                self._guardar_configuracion()
                return items[0]['id']
        except Exception as e:
            print(f'Error buscando carpeta: {e}')
        
        # Crear la carpeta si no existe
        try:
            metadata = {
                'name': nombre,
                'mimeType': 'application/vnd.google-apps.folder'
            }
            
            carpeta = self.service.files().create(body=metadata, fields='id').execute()
            self.carpeta_id = carpeta['id']
            self._guardar_configuracion()
            print(f'Carpeta creada: {nombre} (ID: {carpeta["id"]})')
            return carpeta['id']
        except Exception as e:
            print(f'Error al crear carpeta: {e}')
            return None
    
    def subir_excel_drive(self, ruta_local, actualizar=False):
        """
        Sube o actualiza el Excel en Drive
        """
        if not self.service:
            print('Error: No has iniciado sesiÃ³n.')
            return None
        
        if not os.path.exists(ruta_local):
            print(f'Error: El archivo {ruta_local} no existe.')
            return None
        
        # Asegurar que existe la carpeta
        if not self.carpeta_id:
            self.crear_o_obtener_carpeta()
        
        try:
            # Verificar que el archivo no estÃ© abierto por otro proceso
            import psutil
            archivo_abierto = False
            for proc in psutil.process_iter(['pid', 'name', 'open_files']):
                try:
                    for file in proc.info['open_files'] or []:
                        if ruta_local in file.path:
                            archivo_abierto = True
                            break
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
            
            if archivo_abierto:
                print('Advertencia: El archivo parece estar abierto por otro programa.')
                print('Esperando 2 segundos...')
                time.sleep(2)
            
            # Subir archivo usando mÃ©todo simple (no resumable para archivos pequeÃ±os)
            file_metadata = {
                'name': self.nombre_archivo,
                'parents': [self.carpeta_id] if self.carpeta_id else []
            }
            
            # Usar media body simple para evitar problemas de corrupciÃ³n
            media = MediaFileUpload(
                ruta_local,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                resumable=False  # No usar resumable para archivos pequeÃ±os
            )
            
            if actualizar and self.archivo_excel_id:
                # Actualizar archivo existente
                # Primero eliminar el anterior para evitar conflictos de versiÃ³n
                try:
                    file = self.service.files().update(
                        fileId=self.archivo_excel_id,
                        media_body=media,
                        fields='id, name, mimeType, size'
                    ).execute()
                    
                    print(f'Archivo actualizado exitosamente: {file["name"]}')
                    print(f'ID: {file["id"]}')
                    print(f'Size: {file.get("size", "unknown")} bytes')
                    return file['id']
                    
                except Exception as e:
                    print(f'Error actualizando archivo: {e}')
                    print('Intentando crear nuevo archivo...')
                    actualizar = False
            
            if not actualizar or not self.archivo_excel_id:
                # Crear nuevo archivo
                file = self.service.files().create(
                    body=file_metadata,
                    media_body=media,
                    fields='id, name, mimeType, size'
                ).execute()
                
                self.archivo_excel_id = file['id']
                self._guardar_configuracion()
                
                print(f'Archivo creado exitosamente: {file["name"]}')
                print(f'ID: {file["id"]}')
                print(f'Size: {file.get("size", "unknown")} bytes')
                return file['id']
                
        except Exception as e:
            print(f'Error al subir archivo: {e}')
            import traceback
            traceback.print_exc()
            return None
    
    def descargar_excel_drive(self, ruta_destino=None):
        """Descarga el Excel desde Drive"""
        if not self.service:
            print('Error: No has iniciado sesiÃ³n.')
            return False
        
        if not self.archivo_excel_id:
            print('Error: No hay archivo configurado.')
            return False
        
        if ruta_destino is None:
            import tempfile
            ruta_destino = os.path.join(tempfile.gettempdir(), self.nombre_archivo)
        
        try:
            request = self.service.files().get_media(fileId=self.archivo_excel_id)
            
            with io.FileIO(ruta_destino, 'wb') as f:
                downloader = MediaIoBaseDownload(f, request)
                done = False
                while done is False:
                    status, done = downloader.next_chunk()
                    if status:
                        print(f'Descargando... {int(status.progress() * 100)}%')
            
            print(f'Archivo descargado: {ruta_destino}')
            return ruta_destino
        except Exception as e:
            print(f'Error al descargar archivo: {e}')
            return None
    
    def verificar_excel_drive(self):
        """Verifica si el archivo existe en Drive"""
        if not self.service or not self.archivo_excel_id:
            return False
        
        try:
            file = self.service.files().get(fileId=self.archivo_excel_id, fields='id, name, mimeType').execute()
            print(f'Archivo verificado en Drive: {file["name"]}')
            return True
        except Exception as e:
            print(f'Archivo no encontrado en Drive: {e}')
            self.archivo_excel_id = ''
            self._guardar_configuracion()
            return False
    
    def obtener_enlace_compartido(self):
        """Obtiene el enlace para compartir el archivo"""
        if not self.service or not self.archivo_excel_id:
            print('Error: No hay archivo para compartir.')
            return None
        
        try:
            # Hacer el archivo compartible
            permission = {
                'type': 'anyone',
                'role': 'reader'
            }
            
            self.service.permissions().create(
                fileId=self.archivo_excel_id,
                body=permission
            ).execute()
            
            enlace = f'https://drive.google.com/file/d/{self.archivo_excel_id}/view?usp=sharing'
            print(f'Enlace de acceso: {enlace}')
            return enlace
        except Exception as e:
            print(f'Error al crear enlace: {e}')
            return None

def _resolver_mes_objetivo(month_mode='actual'):
    """Resolver el mes objetivo: actual, siguiente o YYYY-MM."""
    modo = str(month_mode or 'actual').strip().lower()
    ahora = datetime.now()
    anio_actual = ahora.year
    mes_actual = ahora.month

    if modo in ('actual', 'mes_actual', 'current'):
        objetivo = datetime(anio_actual, mes_actual, 1)
        modo_normalizado = 'actual'
    elif modo in ('siguiente', 'mes_siguiente', 'next', 'proximo'):
        if mes_actual == 12:
            objetivo = datetime(anio_actual + 1, 1, 1)
        else:
            objetivo = datetime(anio_actual, mes_actual + 1, 1)
        modo_normalizado = 'siguiente'
    elif re.match(r'^\d{4}-\d{2}$', modo):
        anio = int(modo[:4])
        mes = int(modo[5:7])
        if mes < 1 or mes > 12:
            raise ValueError('Mes invalido en month_mode. Usa YYYY-MM con MM entre 01 y 12.')
        objetivo = datetime(anio, mes, 1)
        modo_normalizado = modo
    else:
        raise ValueError('month_mode invalido. Usa "actual", "siguiente" o "YYYY-MM".')

    meses = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    mes_nombre = meses[objetivo.month - 1]
    anio = objetivo.year
    clave_mes = f'{anio:04d}-{objetivo.month:02d}'
    hoja_objetivo = f'{mes_nombre} {anio}'
    return {
        'month_mode': modo_normalizado,
        'mes_nombre': mes_nombre,
        'anio': anio,
        'clave_mes': clave_mes,
        'hoja_objetivo': hoja_objetivo,
    }


def _calcular_ingresos_extra_mes(config, clave_mes):
    historial = config.get('historial_saldos', {}) if isinstance(config, dict) else {}
    registro_mes = {}
    if isinstance(historial, dict):
        saldos_mensuales = historial.get('saldos_mensuales', {})
        if isinstance(saldos_mensuales, dict):
            registro_mes = saldos_mensuales.get(clave_mes, {}) or {}
        if not registro_mes and isinstance(historial.get(clave_mes), dict):
            registro_mes = historial.get(clave_mes, {})

    ingresos_raw = registro_mes.get('ingresos_extra', []) if isinstance(registro_mes, dict) else []
    total = 0.0
    count = 0
    if isinstance(ingresos_raw, list):
        for item in ingresos_raw:
            if not isinstance(item, dict):
                continue
            try:
                valor = float(item.get('valor', 0) or 0)
            except Exception:
                valor = 0.0
            if valor > 0:
                total += valor
                count += 1
    return total, count


def sincronizar_con_drive(config_path='config/configuracion.json', month_mode='actual'):
    """Sincronizar Excel con Drive creando/actualizando la hoja del mes objetivo."""
    try:
        from excel_mensual import GeneradorExcelMensual
    except ModuleNotFoundError:
        from src.excel_mensual import GeneradorExcelMensual
    import openpyxl

    print('=' * 60)
    print('SINCRONIZACION CON GOOGLE DRIVE')
    print('=' * 60)

    target = _resolver_mes_objetivo(month_mode)
    mes_nombre = target['mes_nombre']
    anio_objetivo = target['anio']
    hoja_objetivo = target['hoja_objetivo']
    clave_mes = target['clave_mes']
    month_mode = target['month_mode']

    drive = GoogleDriveManager(config_path)
    excel_gen = GeneradorExcelMensual(config_path)

    if not drive.autenticar():
        msg = 'No se pudo autenticar con Drive'
        print(f'Error: {msg}')
        return {'success': False, 'message': msg}

    carpeta_id = drive.crear_o_obtener_carpeta()
    if not carpeta_id:
        msg = 'No se pudo crear/obtener carpeta en Drive'
        print(f'Error: {msg}')
        return {'success': False, 'message': msg}

    wb = None
    archivo_existente = False
    hoja_ya_existia = False

    if drive.archivo_excel_id and drive.verificar_excel_drive():
        print('')
        print('Archivo existente encontrado en Drive')
        print('Descargando para actualizar...')
        ruta_temp = drive.descargar_excel_drive()
        if ruta_temp:
            wb = openpyxl.load_workbook(ruta_temp)
            archivo_existente = True
            hoja_ya_existia = hoja_objetivo in wb.sheetnames or mes_nombre in wb.sheetnames

    if wb is None:
        print('')
        print('Creando nuevo archivo Excel base...')
        wb = excel_gen.crear_excel_nuevo()
        hoja_ya_existia = hoja_objetivo in wb.sheetnames or mes_nombre in wb.sheetnames

    excel_gen.crear_o_actualizar_hoja_mes(wb, mes_nombre, anio_objetivo)
    print(f'Hoja objetivo preparada: {hoja_objetivo}')
    ruta_nueva = excel_gen.guardar_excel_temporal(wb)

    file_id = drive.subir_excel_drive(ruta_nueva, actualizar=archivo_existente)
    if not file_id:
        msg = 'No se pudo subir el archivo a Drive'
        print(f'Error: {msg}')
        return {'success': False, 'message': msg}

    enlace = drive.obtener_enlace_compartido()
    ingresos_extra_total, ingresos_extra_count = _calcular_ingresos_extra_mes(excel_gen.config, clave_mes)

    print('')
    print('=' * 60)
    print('SINCRONIZACION COMPLETADA')
    print('=' * 60)
    if enlace:
        print('')
        print(f'Enlace para acceder: {enlace}')
    print('')
    print('El archivo se guardo en la carpeta: ControlDeGastos')
    print(f'Con la hoja: {hoja_objetivo}')

    return {
        'success': True,
        'message': 'Sincronizacion completada',
        'month_mode': month_mode,
        'hoja_objetivo': hoja_objetivo,
        'mes_clave': clave_mes,
        'hoja_creada': not hoja_ya_existia,
        'archivo_existente': archivo_existente,
        'file_id': file_id,
        'enlace': enlace,
        'ingresos_extra_total': ingresos_extra_total,
        'ingresos_extra_count': ingresos_extra_count,
    }


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Sincronizar Excel con Google Drive')
    parser.add_argument(
        '--month-mode',
        default='actual',
        help='Mes objetivo: actual, siguiente o YYYY-MM',
    )
    args = parser.parse_args()
    result = sincronizar_con_drive(month_mode=args.month_mode)
    if not result.get('success'):
        raise SystemExit(1)
