import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
import json

class GeneradorExcelGastos:
    
    def __init__(self, config_path='config/configuracion.json'):
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = json.load(f)
        
        self.colores = {
            'encabezado': '366092',
            'subencabezado': '4472C4',
            'ingreso': 'C6EFCE',
            'gasto': 'FFC7CE',
            'total': 'FFEB9C',
            'neutro': 'D9D9D9'
        }
        
        self.bordes = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def crear_libro_nuevo(self, nombre_mes=None):
        if nombre_mes is None:
            nombre_mes = datetime.now().strftime('%Y-%m')
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        ws_resumen = wb.create_sheet('Resumen')
        self._crear_hoja_resumen(ws_resumen)
        
        ws_detalle = wb.create_sheet('Detalle')
        self._crear_hoja_detalle(ws_detalle)
        
        ws_historico = wb.create_sheet('Histórico')
        self._crear_hoja_historico(ws_historico)
        
        ws_dashboard = wb.create_sheet('Dashboard')
        self._crear_hoja_dashboard(ws_dashboard)
        
        return wb, nombre_mes
    
    def _crear_hoja_resumen(self, ws):
        ws['A1'] = 'CONTROL DE GASTOS MENSUALES'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colores['encabezado'], end_color=self.colores['encabezado'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        ws['A3'] = 'INGRESOS'
        ws['A3'].font = Font(size=12, bold=True, color='FFFFFF')
        ws['A3'].fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
        ws.merge_cells('A3:D3')
        
        ws['A4'] = 'Concepto'
        ws['B4'] = 'Monto (COP)'
        ws['C4'] = '% del Total'
        ws['D4'] = 'Observaciones'
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}4'].font = Font(bold=True)
            ws[f'{col}4'].fill = PatternFill(start_color=self.colores['subencabezado'], end_color=self.colores['subencabezado'], fill_type='solid')
            ws[f'{col}4'].border = self.bordes
        
        ws['A5'] = 'Sueldo Mensual'
        ws['B5'] = self.config['sueldo']['valor_fijo']
        ws['B5'].number_format = '$#,##0'
        ws['C5'] = '100%'
        ws['C5'].number_format = '0%'
        
        ws['A6'] = 'TOTAL INGRESOS'
        ws['A6'].font = Font(bold=True)
        ws['A6'].fill = PatternFill(start_color=self.colores['ingreso'], end_color=self.colores['ingreso'], fill_type='solid')
        ws['B6'] = f'=B5'
        ws['B6'].number_format = '$#,##0'
        ws['B6'].font = Font(bold=True)
        ws['B6'].fill = PatternFill(start_color=self.colores['ingreso'], end_color=self.colores['ingreso'], fill_type='solid')
        ws['C6'] = '100%'
        ws['C6'].number_format = '0%'
        ws['C6'].fill = PatternFill(start_color=self.colores['ingreso'], end_color=self.colores['ingreso'], fill_type='solid')
        
        for row in range(5, 7):
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = self.bordes
        
        ws['A8'] = 'GASTOS FIJOS'
        ws['A8'].font = Font(size=12, bold=True, color='FFFFFF')
        ws['A8'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        ws.merge_cells('A8:D8')
        
        ws['A9'] = 'Concepto'
        ws['B9'] = 'Monto (COP)'
        ws['C9'] = '% Ingresos'
        ws['D9'] = 'Día de Cargo'
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}9'].font = Font(bold=True)
            ws[f'{col}9'].fill = PatternFill(start_color=self.colores['subencabezado'], end_color=self.colores['subencabezado'], fill_type='solid')
            ws[f'{col}9'].border = self.bordes
        
        fila = 10
        gastos_fijos = self.config['gastos_fijos']
        
        for concepto, datos in gastos_fijos.items():
            nombre_concepto = concepto.replace('_', ' ').title()
            ws[f'A{fila}'] = nombre_concepto
            ws[f'B{fila}'] = datos['valor']
            ws[f'B{fila}'].number_format = '$#,##0'
            ws[f'C{fila}'] = f'=B{fila}/$B$6'
            ws[f'C{fila}'].number_format = '0.00%'
            
            if 'dia_cargo' in datos:
                ws[f'D{fila}'] = f"Día {datos['dia_cargo']}"
            elif 'frecuencia' in datos:
                ws[f'D{fila}'] = datos['frecuencia'].title()
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{fila}'].border = self.bordes
            
            fila += 1
        
        ws[f'A{fila}'] = 'TOTAL GASTOS FIJOS'
        ws[f'A{fila}'].font = Font(bold=True)
        ws[f'A{fila}'].fill = PatternFill(start_color=self.colores['gasto'], end_color=self.colores['gasto'], fill_type='solid')
        ws[f'B{fila}'] = f'=SUM(B10:B{fila-1})'
        ws[f'B{fila}'].number_format = '$#,##0'
        ws[f'B{fila}'].font = Font(bold=True)
        ws[f'B{fila}'].fill = PatternFill(start_color=self.colores['gasto'], end_color=self.colores['gasto'], fill_type='solid')
        ws[f'C{fila}'] = f'=B{fila}/$B$6'
        ws[f'C{fila}'].number_format = '0.00%'
        ws[f'C{fila}'].fill = PatternFill(start_color=self.colores['gasto'], end_color=self.colores['gasto'], fill_type='solid')
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{fila}'].border = self.bordes
        
        fila += 2
        
        ws[f'A{fila}'] = 'GASTOS VARIABLES'
        ws[f'A{fila}'].font = Font(size=12, bold=True, color='FFFFFF')
        ws[f'A{fila}'].fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
        ws.merge_cells(f'A{fila}:D{fila}')
        fila += 1
        
        ws[f'A{fila}'] = 'Concepto'
        ws[f'B{fila}'] = 'Monto (COP)'
        ws[f'C{fila}'] = '% Ingresos'
        ws[f'D{fila}'] = 'Fecha'
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{fila}'].font = Font(bold=True)
            ws[f'{col}{fila}'].fill = PatternFill(start_color=self.colores['subencabezado'], end_color=self.colores['subencabezado'], fill_type='solid')
            ws[f'{col}{fila}'].border = self.bordes
        
        fila += 1
        inicio_gastos_variables = fila
        
        ws[f'A{fila}'] = 'Gastos Generales'
        ws[f'B{fila}'] = 0
        ws[f'B{fila}'].number_format = '$#,##0'
        ws[f'C{fila}'] = f'=B{fila}/$B$6'
        ws[f'C{fila}'].number_format = '0.00%'
        ws[f'D{fila}'] = 'Variable'
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{fila}'].border = self.bordes
        
        fila += 1
        
        ws[f'A{fila}'] = 'TOTAL GASTOS VARIABLES'
        ws[f'A{fila}'].font = Font(bold=True)
        ws[f'A{fila}'].fill = PatternFill(start_color=self.colores['gasto'], end_color=self.colores['gasto'], fill_type='solid')
        ws[f'B{fila}'] = f'=SUM(B{inicio_gastos_variables}:B{fila-1})'
        ws[f'B{fila}'].number_format = '$#,##0'
        ws[f'B{fila}'].font = Font(bold=True)
        ws[f'B{fila}'].fill = PatternFill(start_color=self.colores['gasto'], end_color=self.colores['gasto'], fill_type='solid')
        ws[f'C{fila}'] = f'=B{fila}/$B$6'
        ws[f'C{fila}'].number_format = '0.00%'
        ws[f'C{fila}'].fill = PatternFill(start_color=self.colores['gasto'], end_color=self.colores['gasto'], fill_type='solid')
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{fila}'].border = self.bordes
        
        total_gastos_fijos = fila - 3
        total_gastos_variables = fila
        
        fila += 2
        
        ws[f'A{fila}'] = 'RESUMEN FINAL'
        ws[f'A{fila}'].font = Font(size=12, bold=True, color='FFFFFF')
        ws[f'A{fila}'].fill = PatternFill(start_color=self.colores['encabezado'], end_color=self.colores['encabezado'], fill_type='solid')
        ws.merge_cells(f'A{fila}:D{fila}')
        fila += 1
        
        ws[f'A{fila}'] = 'Total Ingresos'
        ws[f'B{fila}'] = f'=B6'
        ws[f'B{fila}'].number_format = '$#,##0'
        ws[f'A{fila}'].font = Font(bold=True)
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{fila}'].border = self.bordes
        
        fila += 1
        
        ws[f'A{fila}'] = 'Total Gastos Fijos'
        ws[f'B{fila}'] = f'=B{total_gastos_fijos}'
        ws[f'B{fila}'].number_format = '$#,##0'
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{fila}'].border = self.bordes
        
        fila += 1
        
        ws[f'A{fila}'] = 'Total Gastos Variables'
        ws[f'B{fila}'] = f'=B{total_gastos_variables}'
        ws[f'B{fila}'].number_format = '$#,##0'
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{fila}'].border = self.bordes
        
        fila += 1
        
        ws[f'A{fila}'] = 'TOTAL GASTOS'
        ws[f'A{fila}'].font = Font(bold=True)
        ws[f'A{fila}'].fill = PatternFill(start_color=self.colores['gasto'], end_color=self.colores['gasto'], fill_type='solid')
        ws[f'B{fila}'] = f'=B{fila-2}+B{fila-1}'
        ws[f'B{fila}'].number_format = '$#,##0'
        ws[f'B{fila}'].font = Font(bold=True)
        ws[f'B{fila}'].fill = PatternFill(start_color=self.colores['gasto'], end_color=self.colores['gasto'], fill_type='solid')
        ws[f'C{fila}'] = f'=B{fila}/B6'
        ws[f'C{fila}'].number_format = '0.00%'
        ws[f'C{fila}'].fill = PatternFill(start_color=self.colores['gasto'], end_color=self.colores['gasto'], fill_type='solid')
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{fila}'].border = self.bordes
        
        fila += 1
        
        ws[f'A{fila}'] = 'AHORRO / SOBRANTE'
        ws[f'A{fila}'].font = Font(bold=True, size=11)
        ws[f'A{fila}'].fill = PatternFill(start_color=self.colores['total'], end_color=self.colores['total'], fill_type='solid')
        ws[f'B{fila}'] = f'=B{fila-4}-B{fila-1}'
        ws[f'B{fila}'].number_format = '$#,##0'
        ws[f'B{fila}'].font = Font(bold=True, size=11)
        ws[f'B{fila}'].fill = PatternFill(start_color=self.colores['total'], end_color=self.colores['total'], fill_type='solid')
        ws[f'C{fila}'] = f'=B{fila}/B{fila-4}'
        ws[f'C{fila}'].number_format = '0.00%'
        ws[f'C{fila}'].fill = PatternFill(start_color=self.colores['total'], end_color=self.colores['total'], fill_type='solid')
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{fila}'].border = self.bordes
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
    
    def _crear_hoja_detalle(self, ws):
        ws['A1'] = 'REGISTRO DETALLADO DE GASTOS'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colores['encabezado'], end_color=self.colores['encabezado'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        headers = ['Fecha', 'Categoría', 'Concepto', 'Monto (COP)', 'Método de Pago', 'Notas']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colores['subencabezado'], end_color=self.colores['subencabezado'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.bordes
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 40
        
        ws.freeze_panes = 'A4'
    
    def _crear_hoja_historico(self, ws):
        ws['A1'] = 'HISTÓRICO MENSUAL'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colores['encabezado'], end_color=self.colores['encabezado'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 30
        
        headers = ['Mes', 'Ingresos', 'Gastos Fijos', 'Gastos Variables', 'Total Gastos', 'Ahorro', '% Ahorro', 'Observaciones']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colores['subencabezado'], end_color=self.colores['subencabezado'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.bordes
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 40
        
        ws.freeze_panes = 'A4'
    
    def _crear_hoja_dashboard(self, ws):
        ws['A1'] = 'DASHBOARD - ANÁLISIS VISUAL'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colores['encabezado'], end_color=self.colores['encabezado'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        ws['A3'] = 'Resumen del Mes'
        ws['A3'].font = Font(size=12, bold=True)
        
        ws['A5'] = 'Este dashboard mostrará gráficos automáticos basados en los datos del mes.'
        ws['A5'].font = Font(italic=True)
        ws['A6'] = 'Los gráficos se generan automáticamente al actualizar los datos.'
        ws['A6'].font = Font(italic=True)
        
        ws.column_dimensions['A'].width = 60
    
    def guardar_excel(self, wb, nombre_archivo):
        ruta = f'excel_templates/{nombre_archivo}.xlsx'
        wb.save(ruta)
        return ruta
    
    def crear_plantilla_inicial(self):
        wb, nombre_mes = self.crear_libro_nuevo()
        ruta = self.guardar_excel(wb, 'ControlDeGastos')
        return ruta

if __name__ == '__main__':
    generador = GeneradorExcelGastos()
    ruta = generador.crear_plantilla_inicial()
    print(f'Excel creado exitosamente en: {ruta}')
